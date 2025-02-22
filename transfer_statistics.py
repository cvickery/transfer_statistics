#! /usr/local/bin/python3
"""Count transferred courses and how many credits transfer as blanket credits (bkcr).

This code is an alternative to count_transfers.py, which simply ranks all courses transferred by
their frequency of transfer. Here, courses are still ranked by frequency, but also by the percentage
of credits that transfer as “real” course credits—those credits that have the potential of counting
towards a program's requirements rather than free-electives. We don't count whether real courses
actually do count towards program requirements ... yet. And we don't look at how GenEd requirements
factor into the transfer process ... yet.

This code gathers information about sending-side courses (“src_course”) from our database of CUNY
courses and transfer rules. It then goes through the spreadsheet from a CUNYfirst query that reports
details about all course transfers across CUNY colleges to gather information about how each
src_course transferred to each receiving college (dst_institution).

Everything is ogranized by dst_institution because the resulting workbook is organized as one
spreadsheet per receiving college.

Dictionaries
  src_courses
    Keys    [dst_institution][(src_course_id, src_offer_nbr)]
    Values  SrcCourse(src_institution, course_str, transfer_rules)
  rule_descriptions
    Key     rule_key
    Value   Natural language text
  metadata
    Key     (course_id, offer_nbr)
    Values  Metadata(institution, course_str,
                     is_undergraduate, is_active, is_mesg, is_bkcr, is_unknown)
            .flags_str is textual representation of the five booleans

"""

import csv
import os
import psycopg
import time

from adjustcolwidths import adjust_widths
from transfer_by_subjects import create_sender_subject_workbook
from collections import defaultdict, namedtuple
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Alignment, Font
from pathlib import Path
from psycopg.rows import namedtuple_row
from recordclass import recordclass

DEBUG = os.getenv('DEBUG_TRANSFER_STATISTICS')

metadata = dict()  # Index by (course_id, offer_nbr)


# elapsed()
# -------------------------------------------------------------------------------------------------
def elapsed(since: float):
  """Return the hours, minutes, and seconds since “since” seconds ago."""
  h, ms = divmod(int(time.time() - since), 3600)
  m, s = divmod(ms, 60)
  return f'{h:02}:{m:02}:{s:02}'


if __name__ == '__main__':

  session_start = time.time()
  log_file = open('transfer_statistics.log', 'w')
  report_file = open(f'./reports/{datetime.today().isoformat()[0:10]}_report.txt', 'w')

  # Check CUNYfirst data is current
  latest_query = None
  query_files = Path('./downloads/').glob('CV_QNS_XFER_STATS*csv')
  for query_file in query_files:
    if latest_query is None:
      latest_query = query_file
      latest_timestamp = query_file.stat().st_mtime
    else:
      this_timestamp = query_file.stat().st_mtime
      if this_timestamp > latest_timestamp:
        latest_query = query_file
        latest_timestamp = this_timestamp
  file_date = f'{time.strftime("%Y-%m-%d", time.localtime(latest_timestamp))}'
  file_age = (datetime.today() - datetime.fromisoformat(file_date)).days
  if file_age > 7:
    exit(f'CUNYfirst info from ({latest_query}) file is {file_age} days old')
  s = '' if file_age == 1 else 's'
  print(f'CUNYfirst query ({latest_query}) is {file_age} day{s} old.')

  # Initialize From Curriculum Database
  # ===============================================================================================

  # A SrcCourse is one that has one or more xfer rules that awards bkcr
  SrcCourse = namedtuple('SrcCourse', 'src_institution, course_str, rules')
  src_courses = defaultdict(dict)  # Index by [dst_institution][src_course_id, src_offer_nbr]

  with psycopg.connect('dbname=cuny_curriculum') as conn:
    with conn.cursor(row_factory=namedtuple_row) as cursor:

      # Be sure rule descriptions are up to date
      cursor.execute("select update_date from updates where table_name = 'rule_descriptions'")
      update_date = cursor.fetchone().update_date
      days = (datetime.today() - datetime.fromisoformat(update_date)).days
      if days > 7:
        exit(f'rule_descriptions have not been updated in {days} days.')
      s = '' if days == 1 else 's'
      print(f'Rule descriptions were updated {days} day{s} ago.')

      print('Collect Transfer Rules')

      cursor.execute("""
      select src.course_id, src.offer_nbr, src.discipline, src.catalog_number,
             rules.source_institution,
             rules.destination_institution,
             string_agg(rule_key, ' ') as rules
      from source_courses src, transfer_rules rules, destination_courses dst
      where (src.course_id, src.offer_nbr, rules.destination_institution) in
            (select s.course_id, s.offer_nbr, r.destination_institution
               from source_courses s, transfer_rules r, destination_courses d
               where s.rule_id = r.id
                 and d.rule_id = r.id
                 and (d.is_bkcr or d.is_mesg)
               group by s.course_id, s.offer_nbr, r.destination_institution)
        and src.rule_id = rules.id
        and dst.rule_id = rules.id
      group by src.course_id, src.offer_nbr, src.discipline, src.catalog_number,
               rules.source_institution, rules.destination_institution
      """)

      for row in cursor:
        course_str = f'{row.discipline.strip()} {row.catalog_number.strip()}'
        src_key = (row.course_id, row.offer_nbr)
        dest = row.destination_institution
        src_courses[dest][src_key] = SrcCourse._make([row.source_institution,
                                                      course_str,
                                                      row.rules.split()])
      print(f'  {cursor.rowcount:10,} Sending Courses\t{elapsed(session_start)}')

      # Cache all rule decriptions, previously stored in the cuny_curriculum db.
      rule_descriptions = defaultdict(str)
      cursor.execute("""
      select rule_key, description
      from rule_descriptions
      """)
      for row in cursor:
        rule_descriptions[row.rule_key] = row.description
      print(f'  {len(rule_descriptions):10,} Rule Descriptions\t{elapsed(session_start)}')

      # Cache metadata for all cuny courses, and credits for real courses. Note: this info is
      # used only for receiving courses.
      meta_start = time.time()
      Metadata = namedtuple('Metadata', 'institution course_str '
                                        'is_ugrad is_active is_mesg is_bkcr is_unknown')

      def _flags_str(self):
        """ String giving status of “interesting” settings of the Metadata boolean values.
            Undergraduate-active-real courses will return the empty string.
        """
        return_str = ''
        if not self.is_ugrad:
          return_str += 'G'
        if not self.is_active:
          return_str += 'I'
        if self.is_mesg:
          return_str += 'M'
        if self.is_bkcr:
          return_str += 'B'
        if self.is_unknown:
          return_str += '?'
        return return_str
      setattr(Metadata, 'flags', _flags_str)

      real_credit_courses = set()  # Members are (course_id, offer_nbr)

      cursor.execute("""
      select course_id, offer_nbr, institution, discipline, catalog_number,
             career ~* '^U' as is_ugrad,
             course_status = 'A' as is_active,
             designation in ('MNL', 'MLA') as is_mesg,
             attributes ~* 'bkcr' as is_bkcr
      from cuny_courses
      """)
      for row in cursor:
        course_str = f'{row.discipline.strip()} {row.catalog_number.strip()}'
        metadata[(row.course_id, row.offer_nbr)] = Metadata._make([row.institution,
                                                                   course_str,
                                                                   row.is_ugrad,
                                                                   row.is_active,
                                                                   row.is_mesg,
                                                                   row.is_bkcr,
                                                                   False])
        if not (row.is_mesg or row.is_bkcr):
          real_credit_courses.add((row.course_id, row.offer_nbr))

      print(f'  {len(real_credit_courses):10,} Real-credit courses', file=report_file)
      print(f'  {len(metadata):10,} All courses\t{elapsed(session_start)}')

  # Process latest transfer evaluations query file.
  # =============================================================================================

  print(f'\nGenerate Statistics {latest_query.name[0:-4].strip("-0123456789")} {file_date}'
        )
  print(f'\nTransfer Statistics {latest_query.name[0:-4].strip("-0123456789")} '
        f'{time.strftime("%Y-%m-%d", time.localtime(latest_timestamp))}', file=report_file)

  print(f'{len(open(latest_query, errors="replace").readlines()):11,} Transfer records',
        file=report_file)
  lookup_start = time.time()

  # XferCounts
  # ----------
  """ Indexed by [dst_institution]
      How many records; how many ignored
  """
  XferCounts = recordclass('XferCounts', 'total not_bkcr')

  def xfer_counts_factory():
    return XferCounts._make([0, 0])

  xfer_counts = defaultdict(xfer_counts_factory)
  zero_units_taken = 0  # This is a src_institution value

  # XferStats
  # ---------
  """ Indexed by [dst_institution][src_course]
  """
  DstCourse = recordclass('DstCourse', 'count flags')
  XferStats = recordclass('XferStats', 'num_evaluations students_set '
                          'units_taken real_credits bkcr_credits courses, rules')

  def dst_course_factory():
    return DstCourse._make([0, ''])

  def xfer_stats_maker():
    return XferStats._make((0, set(), 0.0, 0.0, 0.0, defaultdict(dst_course_factory), ''))

  def xfer_stats_factory():
    return defaultdict(xfer_stats_maker)

  xfer_stats = defaultdict(xfer_stats_factory)
  first_post = datetime.today()
  last_post = datetime(1900, 1, 1)
  first_term = 9999
  last_term = 0
  with open(latest_query, newline='', errors='replace') as query_file:
    reader = csv.reader(query_file)
    for line in reader:
      print(f'\r{reader.line_num:,}', end='')
      if reader.line_num == 1:
        Row = namedtuple('Row', [c.lower().replace(' ', '_') for c in line])
      else:
        row = Row._make(line)

        try:
          articulation_term = int(row.articulation_term)
          if articulation_term > last_term:
            last_term = articulation_term
          if articulation_term < first_term:
            first_term = articulation_term
        except ValueError:
          print(f'Ignoring articulation_term on line {reader.line_num:,}: '
                f'“{row.articulation_term}”', file=log_file)

        try:
          m, d, y = row.posted_date.split('/')
          posted_date = datetime(int(y), int(m), int(d))
          if posted_date > last_post:
            last_post = posted_date
          if posted_date < first_post:
            first_post = posted_date
        except ValueError:
          print(f'Ignoring posted_date on line {reader.line_num:,}: '
                f'“{row.posted_date}”', file=log_file)

        # Ignore how non-credit courses transfer. They are presumably used for things like
        # Pathways exemptions, and not relevant for our analysis of which credit-bearing courses
        # fail to transfer as real courses.
        src_units_taken = float(row.units_taken)
        if src_units_taken == 0.0:
          zero_units_taken += 1
          continue

        src_course = (int(row.src_course_id), int(row.src_offer_nbr))
        dst_institution = row.dst_institution
        dst_course = (int(row.dst_course_id), int(row.dst_offer_nbr))

        xfer_counts[dst_institution].total += 1
        if src_course not in src_courses[dst_institution].keys():
          # Not a course of interest: no blanket credit rules for this course (although I guess
          # bkcr could be awarded anyway)
          xfer_counts[dst_institution].not_bkcr += 1
          continue

        dst_rule_descriptions = [f'{rule_descriptions[rule_key]}|{rule_key}'
                                 for rule_key
                                 in src_courses[dst_institution][src_course].rules]

        # Log cases where the subject and catalog number don't match current cuny_courses info.
        # -------------------------------------------------------------------------------------
        src_course_str = f'{row.src_subject.strip()} {row.src_catalog_nbr.strip()}'
        if src_course_str != src_courses[dst_institution][src_course].course_str:
          print(f'Catalog course str ({src_courses[dst_institution][src_course].course_str}) '
                f'NE src course str ({src_course_str}))', file=log_file)

        # For each source course, count the number of times it was transferred, how many different
        # students were involved (in case of re-evaluations), the total number of units taken.
        xfer_stats[dst_institution][src_course].num_evaluations += 1
        xfer_stats[dst_institution][src_course].students_set.add(row.student_id)
        xfer_stats[dst_institution][src_course].units_taken += src_units_taken

        # Transfer outcomes: what destination course was assigned, and what was its nature?
        dst_course_id = int(row.dst_course_id)
        dst_offer_nbr = int(row.dst_offer_nbr)
        dst_course = (dst_course_id, dst_offer_nbr)
        dst_discipline = row.dst_subject.strip()
        dst_catalog_nbr = row.dst_catalog_nbr.strip()
        dst_course_str = f'{dst_discipline} {dst_catalog_nbr}'
        try:
          dst_meta = metadata[dst_course]
        except KeyError:
          # Gotta fake the metadata
          # discipline catalog_number is_ugrad is_active is_mesg is_bkcr, is_unknown
          dst_meta = Metadata._make([dst_institution, dst_course_str,
                                     False, False, False, False, True])

        # Log cases where the subject and catalog number don't match current cuny_courses info.
        # -------------------------------------------------------------------------------------
        if dst_meta.course_str != dst_course_str:
          print(f'Catalog course str ({dst_meta.course_str}) NE dst course str '
                f'({dst_course_str}))', file=log_file)

        dst_units_transferred = float(row.units_transferred)
        if dst_units_transferred > src_units_taken:
          print(f'More received ({dst_units_transferred}) than sent ({src_units_taken}) '
                f'{row.student_id} {row.src_course_id:06}:{row.src_offer_nbr} => '
                f'{row.dst_course_id:06}:{row.dst_offer_nbr}',
                file=log_file)
        if dst_course in real_credit_courses:
          xfer_stats[dst_institution][src_course].real_credits += dst_units_transferred
        else:
          xfer_stats[dst_institution][src_course].bkcr_credits += dst_units_transferred

        xfer_stats[dst_institution][src_course].courses[dst_course_str].count += 1
        xfer_stats[dst_institution][src_course].courses[dst_course_str].flags = dst_meta.flags()
        xfer_stats[dst_institution][src_course].rules = dst_rule_descriptions

  print(f'{zero_units_taken:11,} Zero-credit sending courses ignored', file=report_file)
  print(f'\nTransfer Statistics took {elapsed(lookup_start)}')

  print(f'First Post: {first_post.isoformat()[0:10]}\n Last Post: {last_post.isoformat()[0:10]}',
        file=report_file)
  print(f'First Term: {first_term}\n Last Term: {last_term}', file=report_file)

  print('\nCourses Transferred as Real Courses', file=report_file)
  grand_total = grand_not_bkcr = 0
  for institution in sorted(xfer_counts.keys()):
    total = xfer_counts[institution].total
    grand_total += total
    not_bkcr = xfer_counts[institution].not_bkcr
    grand_not_bkcr += not_bkcr
    percent = 100.0 * not_bkcr / total
    print(f'    {institution[0:3]} {not_bkcr:9,} / {total:<9,} = {percent:5.1f}%', file=report_file)
  percent = 100.0 * grand_not_bkcr / grand_total
  print(f'    ALL {grand_not_bkcr:9,} / {grand_total:<9,} = {percent:5.1f}%', file=report_file)

  print('\nGenerate Report')
  # =============================================================================================
  print('\nSpreadsheet Summary', file=report_file)
  report_start = time.time()

  wb = Workbook()
  # Cell formatting options
  bold = Font(bold=True)

  center_top = NamedStyle('center_top')
  center_top.alignment = Alignment(horizontal='center', vertical='top', wrapText=True)
  center_top.font = bold
  wb.add_named_style(center_top)

  left_top = NamedStyle('left_top')
  left_top.alignment = Alignment(horizontal='left', vertical='top', wrapText=True)
  wb.add_named_style(left_top)

  counter_format = NamedStyle('counter_format')
  counter_format.alignment = Alignment(vertical='top')
  counter_format.number_format = '#,##0'
  wb.add_named_style(counter_format)

  decimal_format = NamedStyle('decimal_format')
  decimal_format.alignment = Alignment(vertical='top')
  decimal_format.number_format = '0.0'
  wb.add_named_style(decimal_format)

  highlighted = Font(bold=True, color='800080')

  headings = ['Sending College', 'Sending Course', 'Students', 'Repeats', 'Sending Cr',
              'Real', 'BKCR', '% Real', 'Receiving Courses', 'Rule Descriptions', 'Rule Keys']

  for dst_institution in sorted(xfer_counts.keys()):
    print(f'\n{dst_institution[0:3]}', file=log_file)
    ws = wb.create_sheet(dst_institution[0:3])
    for col in range(len(headings)):
      ws.cell(1, col + 1, headings[col]).style = 'center_top'

    # Sort dst_institution’s src_course counts
    institution_dict = {key: xfer_stats[dst_institution][key]
                        for key in xfer_stats[dst_institution]}
    row_keys = list(institution_dict.keys())
    row_keys = sorted(row_keys, key=lambda k: institution_dict[k].num_evaluations, reverse=True)
    ws_row_index = 1
    for row_key in row_keys:

      ws_row_index += 1
      src_meta = metadata[row_key]
      ws.cell(ws_row_index, 1, src_meta.institution).style = 'left_top'
      if flags_str := src_meta.flags():
        flags_str = f' [{flags_str}]'
      ws.cell(ws_row_index, 2, f'{src_meta.course_str}{flags_str}').style = 'left_top'

      num_evaluations = institution_dict[row_key].num_evaluations
      num_students = len(institution_dict[row_key].students_set)
      num_reevaluations = (num_evaluations - num_students)
      assert num_reevaluations >= 0
      ws.cell(ws_row_index, 3, num_students).style = 'counter_format'
      ws.cell(ws_row_index, 4, num_reevaluations).style = 'counter_format'

      units_taken = institution_dict[row_key].units_taken / num_evaluations
      real_credits = institution_dict[row_key].real_credits / num_evaluations
      bkcr_credits = institution_dict[row_key].bkcr_credits / num_evaluations
      credits_lost = units_taken - (real_credits + bkcr_credits)
      percent_real = (100.0 * real_credits) / (real_credits + bkcr_credits + credits_lost)
      do_highlight = percent_real < 50.0
      ws.cell(ws_row_index, 5, units_taken).style = 'decimal_format'
      ws.cell(ws_row_index, 6, real_credits).style = 'decimal_format'
      ws.cell(ws_row_index, 7, bkcr_credits).style = 'decimal_format'
      ws.cell(ws_row_index, 8, percent_real).style = 'decimal_format'

      courses_list = []
      for course in institution_dict[row_key].courses:
        flags_str = institution_dict[row_key].courses[course].flags
        if flags_str:
          flags_str = f' [{flags_str}]'
        courses_list.append(f'{course}{flags_str} '
                            f'({institution_dict[row_key].courses[course].count:,})')
      ws.cell(ws_row_index, 9, '\n'.join(courses_list)).style = 'left_top'

      rule_descriptions = []
      rule_keys = []
      for rule in institution_dict[row_key].rules:
        rule_description, rule_key = rule.split('|')
        rule_descriptions.append(rule_description)
        rule_keys.append(rule_key)
      ws.cell(ws_row_index, 10, '\n'.join(rule_descriptions)).style = 'left_top'
      ws.cell(ws_row_index, 11, '\n'.join(rule_keys)).style = 'left_top'

      if do_highlight:
        for col_index in range(1, len(headings) + 1):
          ws.cell(ws_row_index, col_index).font = highlighted
    s = '' if ws_row_index == 1 else 's'
    print(f'{dst_institution} {ws_row_index:6,} row{s}', file=report_file)

  # Clean up

  del wb['Sheet']
  adjust_widths(wb, [8.0, 10.0, 10.0, 10.0, 10.0, 10.0, 10.0, 10.0, 20.0, 150.0, 20.0])
  wb.save(f'./reports/{datetime.today().isoformat()[0:10]}_transfer_statistics.xlsx')

  print('\nReport time\t', elapsed(report_start))
  print('Total time\t', elapsed(session_start))

  log_file.close()
  report_file.close()

  print('\nGenerate Institution Reports')
  for dst_institution in sorted(xfer_counts.keys()):
    print(f'\nProcessing {dst_institution}')
    institution_dict = {key: xfer_stats[dst_institution][key]
                        for key in xfer_stats[dst_institution]}

    # Get workbook for this institution.
    wb = create_sender_subject_workbook(dst_institution, institution_dict)
    try:
      wb.save(f'./reports/{datetime.today().isoformat()[0:10]}_'
              f'{dst_institution}_transfer_statistics.xlsx')
      print(f'Created workbook for {dst_institution}')
    except IndexError:
      print(f'No active sheets for {dst_institution}')
