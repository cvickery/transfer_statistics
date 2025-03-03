#! /usr/local/bin/python3

from adjustcolwidths import adjust_widths
from copy import copy
from openpyxl import load_workbook, Workbook
from pathlib import Path


# Input and output paths
input_file = None
for file in Path('reports').glob('20*transfer_statistics.xlsx'):  # Avoid Excel backups (~20...)
  if not input_file or (file.name > input_file.name):
    input_file = file
if not input_file:
  exit('No YYYY-MM-DD_transfer_statistics.xlsx files found')
file_date = input_file.name[0:10]

output_dir = Path('reports/by_institution')
Path.mkdir(output_dir, exist_ok=True)

# Load the existing workbook
input_wb = load_workbook(input_file)

# Process each sheet (institution names: BAR, BCC, etc.)
for institution_name in input_wb.sheetnames:
  institution_ws = input_wb[institution_name]

  # Dictionary to store department-wise sheets within a workbook per institution
  institution_workbook = Workbook()
  institution_workbook.remove(institution_workbook.active)  # Remove default empty sheet

  # Identify the Department column
  department_col = None
  header_row = 1  # Assuming first row contains headers

  for col in range(1, institution_ws.max_column + 1):
    if institution_ws.cell(header_row, col).value == 'Department':
      department_col = col
      break

  if department_col is None:
    print(f'Skipping {institution_name}: No "Department" column found.')
    continue

  # Read data and distribute into department-specific sheets
  department_sheets = {}

  for row in institution_ws.iter_rows(min_row=2, max_row=institution_ws.max_row, values_only=False):
    department_name = row[department_col - 1].value or 'Unknown'  # Handle empty department names
    department_name = department_name.split('-')[0]

    # Create the department sheet if it doesn't exist
    if department_name not in department_sheets:
      dept_ws = institution_workbook.create_sheet(title=department_name[:31])  # Excel name limit
      department_sheets[department_name] = dept_ws

      # Copy column headers
      new_col_idx = 1
      for col_idx, cell in enumerate(institution_ws[header_row], start=1):
        if col_idx == department_col:
          continue
        new_cell = dept_ws.cell(row=1, column=new_col_idx, value=cell.value)
        new_cell.font = copy(cell.font)
        new_cell.alignment = copy(cell.alignment)
        new_cell.fill = copy(cell.fill)
        new_cell.border = copy(cell.border)
        new_col_idx += 1

    else:
      dept_ws = department_sheets[department_name]

    # Append row to the department sheet
    new_row_idx = dept_ws.max_row + 1
    new_col_idx = 1
    for col_idx, cell in enumerate(row, start=1):
      if col_idx == department_col:
        continue
      new_cell = dept_ws.cell(row=new_row_idx, column=new_col_idx, value=cell.value)
      new_cell.font = copy(cell.font)
      new_cell.alignment = copy(cell.alignment)
      new_cell.fill = copy(cell.fill)
      new_cell.border = copy(cell.border)
      new_cell.number_format = copy(cell.number_format)
      new_col_idx += 1

  # Sort sheets: Admin first, others alphabetically
  sheet_names = institution_workbook.sheetnames
  sorted_sheets = ['Admin'] + sorted([name for name in sheet_names if name != 'Admin'])

  # Create a new workbook to store sheets in sorted order
  sorted_workbook = Workbook()
  sorted_workbook.remove(sorted_workbook.active)  # Remove default blank sheet

  for sheet_name in sorted_sheets:
    if sheet_name in institution_workbook.sheetnames:
      ws_original = institution_workbook[sheet_name]
      ws_new = sorted_workbook.create_sheet(title=sheet_name)

      # Copy cell values and styles
      for row in ws_original.iter_rows():
        for cell in row:
          new_cell = ws_new.cell(row=cell.row, column=cell.column, value=cell.value)
          new_cell.font = copy(cell.font)
          new_cell.alignment = copy(cell.alignment)
          new_cell.fill = copy(cell.fill)
          new_cell.border = copy(cell.border)
          new_cell.number_format = copy(cell.number_format)

  # Adjust column widths for the new workbook
  adjust_widths(sorted_workbook,
                [8.0, 10.0, 10.0, 10.0, 10.0, 10.0, 10.0, 10.0, 20.0, 150.0, 20.0, 100.0])

  # Save the sorted workbook
  output_file = Path(output_dir, f'{file_date}_{institution_name}.xlsx')
  try:
    sorted_workbook.save(output_file)
    print(f'Created {output_file}')
  except IndexError:
    print(f'No departments for {institution_name}')
