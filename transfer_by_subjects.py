#! /usr/local/bin/python3

import sys

from adjustcolwidths import adjust_widths

from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, NamedStyle
from shared_metadata import metadata  # Index by (course_id, offer_nbr)


def get_subject(course_str):
  """Extract subject from course string (e.g., 'ENGL 101' -> 'ENGL')"""
  return course_str.split()[0] if course_str else None


def create_sender_subject_workbook(dst_institution, institution_stats):
  """Create a workbook for a receiving institution with sheets organized by sender and subject"""

  wb = Workbook()

  # Cell formatting options
  bold = Font(bold=True)
  center_top = NamedStyle('center_top')
  center_top.alignment = Alignment(horizontal='center', vertical='top', wrapText=True)
  center_top.font = bold
  wb.add_named_style(center_top)

  left_top = NamedStyle('left_top')
  left_top.alignment = Alignment(horizontal='left', vertical='top', wrapText=True)
  wb.add_named_style(left_top)

  counter_format = NamedStyle('counter_format')
  counter_format.alignment = Alignment(vertical='top')
  counter_format.number_format = '#,##0'
  wb.add_named_style(counter_format)

  decimal_format = NamedStyle('decimal_format')
  decimal_format.alignment = Alignment(vertical='top')
  decimal_format.number_format = '0.0'
  wb.add_named_style(decimal_format)

  highlighted = Font(bold=True, color='800080')

  # First, organize data by sending institution
  senders_dict = defaultdict(lambda: defaultdict(list))

  # Collect courses for this receiving institution
  for src_course, stats in institution_stats.items():
    try:
      sender = metadata[src_course].institution[0:3]
      # Get all receiving course subjects for this transfer
      for dst_course_str in stats.courses:
        subject = get_subject(dst_course_str)
        if subject:
          course_data = {
            'sending_college': sender,
            'sending_course': metadata[src_course].course_str,
            'students': len(stats.students_set),
            'repeats': stats.num_evaluations - len(stats.students_set),
            'sending_cr': stats.units_taken / stats.num_evaluations,
            'real': stats.real_credits / stats.num_evaluations,
            'bkcr': stats.bkcr_credits / stats.num_evaluations,
            'percent_real': (100.0 * stats.real_credits) /
            (stats.real_credits + stats.bkcr_credits +
              (stats.units_taken - (stats.real_credits + stats.bkcr_credits))),
            'receiving_courses': dst_course_str,
            'rule_descriptions': '\n'.join(rule.split('|')[0] for rule in stats.rules)
          }
          senders_dict[sender][subject].append(course_data)
    except KeyError:
      print(f'No metadata for {src_course}', file=sys.stderr)

  # Create sheets with naming convention: SendingCollege_Subject
  for sender in sorted(senders_dict.keys()):
    for subject in sorted(senders_dict[sender].keys()):
      sheet_name = subject.replace('*', 'star')
      try:
        ws = wb.create_sheet(title=sheet_name)
      except ValueError:
        breakpoint()

      # Set up headers
      headers = ['Sending College', 'Sending Course', 'Students', 'Repeats', 'Sending Cr',
                 'Real', 'BKCR', '% Real', 'Receiving Courses', 'Rule Descriptions']
      for col, header in enumerate(headers, 1):
        ws.cell(1, col, header).style = 'center_top'

      # Add data
      courses = senders_dict[sender][subject]
      for row, course_data in enumerate(sorted(courses,
                                               key=lambda x: x['students'],
                                               reverse=True), 2):
        ws.cell(row, 1, course_data['sending_college']).style = 'left_top'
        ws.cell(row, 2, course_data['sending_course']).style = 'left_top'
        ws.cell(row, 3, course_data['students']).style = 'counter_format'
        ws.cell(row, 4, course_data['repeats']).style = 'counter_format'
        ws.cell(row, 5, course_data['sending_cr']).style = 'decimal_format'
        ws.cell(row, 6, course_data['real']).style = 'decimal_format'
        ws.cell(row, 7, course_data['bkcr']).style = 'decimal_format'
        ws.cell(row, 8, course_data['percent_real']).style = 'decimal_format'
        ws.cell(row, 9, course_data['receiving_courses']).style = 'left_top'
        ws.cell(row, 10, course_data['rule_descriptions']).style = 'left_top'

        # Highlight rows with less than 50% real credits
        if course_data['percent_real'] < 50.0:
          for col in range(1, len(headers) + 1):
            ws.cell(row, col).font = highlighted

  # Remove default sheet and adjust column widths
  if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

  adjust_widths(wb, [10.0, 10.0, 10.0, 10.0, 10.0, 10.0, 10.0, 10.0, 20.0, 150.0])
  return wb
